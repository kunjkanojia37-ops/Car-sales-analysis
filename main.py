import pandas as pd
import os
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side

# ==============================
# PATH SETUP
# ==============================

base_dir = os.path.dirname(__file__)
raw_path = os.path.join(base_dir, "data", "raw", "car-sales-input.csv")
processed_dir = os.path.join(base_dir, "data", "processed")

os.makedirs(processed_dir, exist_ok=True)

output_excel = os.path.join(processed_dir, "car-sales-output.xlsx")

# ==============================
# LOAD DATA
# ==============================

df = pd.read_csv(raw_path)

# ==============================
# CALCULATIONS
# ==============================

df["Total Tax"] = df["CGST"] + df["SGST"]
df["Revenue"] = df["Quantity"] * df["Selling price"] + df["Total Tax"] - df["Discount"]
df["Profit"] = df["Revenue"] - df["Cost price"]

# ==============================
# MODEL PROFIT ANALYSIS
# ==============================

profit_by_model = df.groupby("Model")["Profit"].sum().sort_values()

plt.figure(figsize=(10, 6))
ax = profit_by_model.plot(kind="bar", color="orange")
plt.title("Profit by Model")
plt.xlabel("Model")
plt.ylabel("Profit")

for container in ax.containers:
    ax.bar_label(container, fmt='₹%.0f')

plt.tight_layout()
plt.savefig(os.path.join(processed_dir, "profit_by_model.png"), dpi=300)
plt.close()

# ==============================
# MONTHLY ANALYSIS
# ==============================

df["Date"] = pd.to_datetime(df["Date"])
df["Month"] = df["Date"].dt.to_period("M")

monthly = df.groupby("Month")["Revenue"].sum().reset_index()
monthly["MoM Growth %"] = monthly["Revenue"].pct_change().mul(100).round(2).fillna(0)

plt.figure(figsize=(8, 5))
plt.plot(monthly["Month"].astype(str), monthly["MoM Growth %"], marker="o")
plt.title("Month-to-Month Growth")
plt.xticks(rotation=45)

for i, v in enumerate(monthly["MoM Growth %"]):
    plt.text(i, v, f"{v:.0f}%", ha="center", va="bottom")

plt.axhline(0)
plt.tight_layout()
plt.savefig(os.path.join(processed_dir, "monthly_growth.png"), dpi=300)
plt.close()

# ==============================
# CITY PROFIT ANALYSIS
# ==============================

profit_in_city = df.groupby("City")["Profit"].sum().sort_values()

plt.figure(figsize=(10, 6))
ax = profit_in_city.plot(kind="bar", color="brown")
plt.title("Profit in City")
plt.xlabel("City")
plt.ylabel("Profit")

for container in ax.containers:
    ax.bar_label(container, fmt='₹%.0f')

plt.tight_layout()
plt.savefig(os.path.join(processed_dir, "profit_in_city.png"), dpi=300)
plt.close()
# ==============================
# KPI SUMMARY
# ==============================

kpi_data = {
    "Metric": [
        "Total Quantity",
        "Total Revenue",
        "Total Profit",
        "Top Model",
        "Worst Model"
        
    ],
    "Value": [
        df["Quantity"].sum(),
        df["Revenue"].sum(),
        df["Profit"].sum(),
        profit_by_model.idxmax(),
        profit_by_model.idxmin()
    ]
}

kpi_df = pd.DataFrame(kpi_data)

# ==============================
# EXPORT TO EXCEL
# ==============================

with pd.ExcelWriter(output_excel, engine="openpyxl") as writer:
    df.to_excel(writer, sheet_name="Full Data", index=False)
    kpi_df.to_excel(writer, sheet_name="KPI Summary", index=False)

# ==============================
# FORMAT EXCEL (ALL SHEETS)
# ==============================

wb = load_workbook(output_excel)

for sheet in wb.sheetnames:
    ws = wb[sheet]

    center = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for column in ws.columns:
        max_length = 0
        col_letter = column[0].column_letter

        for cell in column:
            cell.alignment = center
            cell.border = border

            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[col_letter].width = max_length + 2

wb.save(output_excel)

print("Project executed successfully!")
